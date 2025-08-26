"""Spreadsheet file analyzer for Excel and CSV files"""

from typing import Dict, Any, List, Optional
from .base import AnalyzerBase
import pandas as pd
import openpyxl
import csv
from pathlib import Path


class SpreadsheetAnalyzer(AnalyzerBase):
    """Analyzer for spreadsheet files (.xlsx, .xls, .csv)"""
    
    def analyze(self, file_path: str) -> Dict[str, Any]:
        """Analyze spreadsheet file and extract information"""
        self.validate_file(file_path)
        
        file_ext = Path(file_path).suffix.lower()
        
        if file_ext == '.csv':
            return self._analyze_csv(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            return self._analyze_excel(file_path)
        else:
            raise ValueError(f"Unsupported file type: {file_ext}")
    
    def extract_text(self, file_path: str) -> str:
        """Extract all text from spreadsheet"""
        self.validate_file(file_path)
        
        file_ext = Path(file_path).suffix.lower()
        all_text = []
        
        if file_ext == '.csv':
            df = pd.read_csv(file_path, encoding='utf-8', on_bad_lines='skip')
            all_text.append("=== CSV Data ===")
            all_text.append(df.to_string())
        
        elif file_ext in ['.xlsx', '.xls']:
            excel_file = pd.ExcelFile(file_path)
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                all_text.append(f"=== Sheet: {sheet_name} ===")
                all_text.append(df.to_string())
        
        return '\n\n'.join(all_text)
    
    def _analyze_csv(self, file_path: str) -> Dict[str, Any]:
        """Analyze CSV file"""
        # Detect encoding
        from ..utils.encoder import EncodingHandler
        encoder = EncodingHandler()
        _, encoding = encoder.read_file_auto(file_path)
        
        # Read CSV
        try:
            df = pd.read_csv(file_path, encoding=encoding, on_bad_lines='skip')
        except:
            df = pd.read_csv(file_path, encoding='utf-8', on_bad_lines='skip')
        
        # Analyze data
        analysis = {
            "file_info": self.get_file_info(file_path),
            "encoding": encoding,
            "sheet_info": {
                "sheets": ["CSV Data"],
                "active_sheet": "CSV Data"
            },
            "data_info": {
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": df.columns.tolist(),
                "data_types": df.dtypes.astype(str).to_dict()
            },
            "statistics": self._generate_statistics(df),
            "sample_data": {
                "head": df.head(10).to_dict(),
                "tail": df.tail(5).to_dict()
            },
            "null_values": df.isnull().sum().to_dict(),
            "unique_values": {col: df[col].nunique() for col in df.columns},
            "summary": self._generate_summary(df)
        }
        
        return analysis
    
    def _analyze_excel(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file"""
        # Load workbook
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        excel_file = pd.ExcelFile(file_path)
        
        sheets_data = []
        all_statistics = {}
        
        # Analyze each sheet
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            sheet_info = {
                "name": sheet_name,
                "rows": len(df),
                "columns": len(df.columns),
                "column_names": df.columns.tolist(),
                "data_types": df.dtypes.astype(str).to_dict(),
                "null_values": df.isnull().sum().to_dict(),
                "sample_data": {
                    "head": df.head(5).to_dict(),
                }
            }
            sheets_data.append(sheet_info)
            all_statistics[sheet_name] = self._generate_statistics(df)
        
        return {
            "file_info": self.get_file_info(file_path),
            "workbook_info": {
                "sheet_count": len(wb.sheetnames),
                "sheet_names": wb.sheetnames,
                "active_sheet": wb.active.title if wb.active else None
            },
            "sheets": sheets_data,
            "statistics": all_statistics,
            "formulas": self._extract_formulas(wb),
            "charts": self._count_charts(wb),
            "summary": self._generate_excel_summary(sheets_data)
        }
    
    def _generate_statistics(self, df: pd.DataFrame) -> Dict[str, Any]:
        """Generate statistics for dataframe"""
        stats = {}
        
        # Numeric columns statistics
        numeric_columns = df.select_dtypes(include=['number']).columns
        if len(numeric_columns) > 0:
            stats['numeric'] = df[numeric_columns].describe().to_dict()
        
        # Text columns info
        text_columns = df.select_dtypes(include=['object']).columns
        if len(text_columns) > 0:
            text_stats = {}
            for col in text_columns:
                text_stats[col] = {
                    'unique': df[col].nunique(),
                    'most_common': df[col].mode().iloc[0] if not df[col].mode().empty else None,
                    'missing': df[col].isnull().sum()
                }
            stats['text'] = text_stats
        
        # Date columns
        date_columns = df.select_dtypes(include=['datetime']).columns
        if len(date_columns) > 0:
            date_stats = {}
            for col in date_columns:
                date_stats[col] = {
                    'min': str(df[col].min()),
                    'max': str(df[col].max()),
                    'missing': df[col].isnull().sum()
                }
            stats['dates'] = date_stats
        
        return stats
    
    def _extract_formulas(self, workbook) -> List[Dict[str, Any]]:
        """Extract formulas from Excel workbook"""
        formulas = []
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula
                        formulas.append({
                            'sheet': sheet.title,
                            'cell': cell.coordinate,
                            'formula': cell.value
                        })
                    if len(formulas) >= 100:  # Limit to first 100 formulas
                        return formulas
        
        return formulas
    
    def _count_charts(self, workbook) -> int:
        """Count charts in Excel workbook"""
        chart_count = 0
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_charts'):
                chart_count += len(sheet._charts)
        return chart_count
    
    def _generate_summary(self, df: pd.DataFrame) -> str:
        """Generate summary of dataframe"""
        summary_parts = []
        summary_parts.append(f"データ: {len(df)}行 × {len(df.columns)}列")
        
        if len(df.columns) > 0:
            summary_parts.append(f"列: {', '.join(df.columns[:5].tolist())}")
            if len(df.columns) > 5:
                summary_parts.append(f"他{len(df.columns)-5}列")
        
        # Check for numeric data
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            summary_parts.append(f"数値列: {len(numeric_cols)}個")
        
        return " | ".join(summary_parts)
    
    def _generate_excel_summary(self, sheets_data: List[Dict[str, Any]]) -> str:
        """Generate summary for Excel file"""
        total_rows = sum(s['rows'] for s in sheets_data)
        total_cols = sum(s['columns'] for s in sheets_data)
        sheet_names = [s['name'] for s in sheets_data[:3]]
        
        summary = f"シート数: {len(sheets_data)}, 総データ: {total_rows}行"
        if sheet_names:
            summary += f" | シート: {', '.join(sheet_names)}"
            if len(sheets_data) > 3:
                summary += f" 他{len(sheets_data)-3}シート"
        
        return summary